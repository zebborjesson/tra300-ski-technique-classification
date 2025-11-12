import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import datetime
from openpyxl import load_workbook
from pathlib import Path

def load_data(file_path):
    """Load data from a CSV file into a pandas DataFrame."""
    return pd.read_csv(file_path)

def load_xslsx_data(file_path, sheet_name=None):
    """Load data from an Excel file into a pandas DataFrame."""
    return pd.read_excel(file_path, sheet_name)

def load_xslsx_data_with_row_color_as_column_values(file_path, sheet_name=None):
    """Load data from an Excel file and include row colors as column values."""
    
    # Load workbook with openpyxl to get formatting
    wb = load_workbook(file_path)
    df_dict = {}
    
    for sheet in wb.sheetnames:
        if sheet_name and sheet not in (sheet_name if isinstance(sheet_name, list) else [sheet_name]):
            continue
            
        ws = wb[sheet]
        
        # Read data with pandas
        df = pd.read_excel(file_path, sheet_name=sheet)
        
        # Extract row colors
        row_colors = []
        for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (after header)
            cell = ws.cell(row=row_idx, column=1)  # Get first cell of each row
            if cell.fill.start_color.index != '00000000':  # Not default (no fill)
                color = cell.fill.start_color.rgb
                if color:
                    row_colors.append(color)
                else:
                    row_colors.append('FFFFFF')  # Default white
            else:
                row_colors.append('FFFFFF')  # Default white
        
        # Pad row_colors list to match DataFrame length
        while len(row_colors) < len(df):
            row_colors.append('FFFFFF')
        
        # Add row colors as a new column
        df['RowColor'] = row_colors[:len(df)]
        df_dict[sheet] = df
    
    return df_dict

def load_gear_distribution_data(file_path, sheet_name=None):
    """Load gear distribution data from an Excel file into a pandas DataFrame."""
    return pd.read_excel(file_path, sheet_name)

def save_df_to_csv(df, file_path):
    """Save a pandas DataFrame to a CSV file."""
    df.to_csv(file_path, index=False)

def keep_rows_of_colour(df, color_column='RowColor', color_value='FFFF0000'):
    """Keep only the rows of a specific color."""
    return df[df[color_column] == color_value]

def merge_two_dataframes_on_time(df1, df2, time_col_1='tt_s', time_col_2='ns1:Time', tolerance=1):
    """Merge two DataFrames based on time columns with a specified tolerance."""
    df1 = df1.copy()
    df2 = df2.copy()
    df1[time_col_1] = pd.to_datetime(df1[time_col_1], unit='s')
    df2[time_col_2] = pd.to_datetime(df2[time_col_2], unit='s')
    
    merged_df = pd.merge_asof(df1.sort_values(time_col_1), 
                              df2.sort_values(time_col_2), 
                              left_on=time_col_1, 
                              right_on=time_col_2, 
                              tolerance=pd.Timedelta(seconds=tolerance), 
                              direction='nearest')
    return merged_df

def add_lap_column(df, time_col='tt_s', gap_threshold=5):
    """Add a lap column that starts at 1 and increments when there's a gap > threshold."""
    df = df.copy()
    # Calculate time differences between consecutive rows
    time_diff = df[time_col].diff()
    # If time column is datetime, convert gap_threshold to Timedelta
    if np.issubdtype(df[time_col].dtype, np.datetime64):
        gap_mask = time_diff > pd.Timedelta(seconds=gap_threshold)
    else:
        gap_mask = time_diff > gap_threshold
    # Create lap column starting at 1, incrementing at each gap
    df['Lap'] = gap_mask.cumsum() + 1
    return df

def summarize_data(df):
    """Print summary statistics of the DataFrame."""
    print("Data Summary:")
    print(df.describe())
    print("\nMissing Values:")
    print(df.isnull().sum())

def plot_histogram(df, column):
    """Plot a histogram of a specified column in the DataFrame."""
    plt.figure(figsize=(10, 6))
    plt.hist(df[column].dropna(), bins=30, color='blue', alpha=0.7)
    plt.title(f'Histogram of {column}')
    plt.xlabel(column)
    plt.ylabel('Frequency')
    plt.grid(True)
    plt.show()

def plot_map_from_lat_and_lon_with_time_colourmap(df, lat_col='latitude_deg', lon_col='longitude_deg', time_col='tt_s'):
    """Plot a scatter map using latitude and longitude columns."""
    plt.figure(figsize=(10, 6))
    scatter = plt.scatter(df[lon_col], df[lat_col], c=df[time_col], cmap='viridis', alpha=0.5)
    plt.title('Geographical Scatter Plot with Time Colourmap')
    plt.xlabel('Longitude')
    plt.ylabel('Latitude')
    plt.grid(True)
    plt.colorbar(scatter, label='Time (s)')
    plt.show()

def plot_lat_lon_difference_map(df1, df2, lat_col_1='latitude_deg', lon_col_1='longitude_deg', lat_col_2='ns1:LatitudeDegrees', lon_col_2='ns1:LongitudeDegrees'):
    """Plot the relative error between two sets of latitude and longitude coordinates in meters."""
    # Calculate differences in degrees
    lat_diff = df1[lat_col_1] - df2[lat_col_2]
    lon_diff = df1[lon_col_1] - df2[lon_col_2]
    
    # Convert to meters using approximate conversion
    # 1 degree latitude ≈ 111,000 meters
    # 1 degree longitude ≈ 111,000 * cos(latitude) meters
    avg_lat = (df1[lat_col_1] + df2[lat_col_2]) / 2
    lat_diff_m = lat_diff * 111000
    lon_diff_m = lon_diff * 111000 * np.cos(np.radians(avg_lat))
    
    # Calculate relative error (Euclidean distance in meters)
    relative_error_m = np.sqrt(lat_diff_m**2 + lon_diff_m**2)
    
    plt.figure(figsize=(12, 8))
    
    # Plot with color representing relative error
    scatter = plt.scatter(df1[lon_col_1], df1[lat_col_1], c=relative_error_m, cmap='Reds', alpha=0.7)
    
    plt.title('Geographical Scatter Plot - Relative Error Between Datasets')
    plt.xlabel('Longitude')
    plt.ylabel('Latitude')
    plt.grid(True)
    plt.colorbar(scatter, label='Relative Error (meters)')
    plt.show()
    
    # Print some statistics about the error
    print(f"Mean relative error: {relative_error_m.mean():.2f} meters")
    print(f"Max relative error: {relative_error_m.max():.2f} meters")
    print(f"Min relative error: {relative_error_m.min():.2f} meters")

def plot_map_from_lat_and_lon_with_lap_colourmap(df, lat_col='latitude_deg', lon_col='longitude_deg'):
    """Plot a scatter map using latitude and longitude columns with lap colourmap."""
    plt.figure(figsize=(10, 6))
    scatter = plt.scatter(df[lon_col], df[lat_col], c=df['Lap'], cmap='viridis', alpha=0.5)
    plt.title('Geographical Scatter Plot with Lap Colourmap')
    plt.xlabel('Longitude')
    plt.ylabel('Latitude')
    plt.grid(True)
    plt.colorbar(scatter, label='Lap')
    plt.show()

def plot_and_compare_lat_lon_maps(df1, df2, lat_col_1='latitude_deg', lon_col_1='longitude_deg', lat_col_2="NS1:LatitudeDegrees", lon_col_2="NS1:LongitudeDegrees", time_col_1='tt_s', time_col_2='ns1:Time'):
    """Plot and compare two geographical scatter maps side by side."""
    fig, axes = plt.subplots(1, 2, figsize=(20, 8))

    scatter1 = axes[0].scatter(df1[lon_col_1], df1[lat_col_1], c=df1[time_col_1], cmap='viridis', alpha=0.5)
    axes[0].set_title('Dataset 1 Geographical Scatter Plot')
    axes[0].set_xlabel('Longitude')
    axes[0].set_ylabel('Latitude')
    axes[0].grid(True)
    fig.colorbar(scatter1, ax=axes[0], label='Time (s)')

    scatter2 = axes[1].scatter(df2[lon_col_2], df2[lat_col_2], c=df2[time_col_2], cmap='viridis', alpha=0.5)
    axes[1].set_title('Dataset 2 Geographical Scatter Plot')
    axes[1].set_xlabel('Longitude')
    axes[1].set_ylabel('Latitude')
    axes[1].grid(True)
    fig.colorbar(scatter2, ax=axes[1], label='Time (s)')

    plt.show()

def plot_lat_lon_map_with_gear_colourmap(df, lat_col='NS1:LatitudeDegrees', lon_col='NS1:LongitudeDegrees', gear_col='Gear'):
    """
    Plot a scatter map using latitude and longitude columns with gear colourmap.
    Uses gear values as categorical labels for proper color mapping.
    """
    # Filter out rows with missing gear data
    df_filtered = df.dropna(subset=[gear_col])
    
    if df_filtered.empty:
        print("No gear data available for plotting.")
        return
    
    # Get unique gear values and create a mapping to integers
    unique_gears = sorted(df_filtered[gear_col].unique())
    gear_to_int = {gear: idx for idx, gear in enumerate(unique_gears)}
    
    # Map gear values to integers for color mapping
    gear_numeric = df_filtered[gear_col].map(gear_to_int)
    
    plt.figure(figsize=(12, 8))
    scatter = plt.scatter(df_filtered[lon_col], df_filtered[lat_col], 
                         c=gear_numeric, cmap='tab20', alpha=0.7, s=20)
    plt.title('Geographical Scatter Plot with Gear Colourmap')
    plt.xlabel('Longitude')
    plt.ylabel('Latitude')
    plt.grid(True, alpha=0.3)
    
    # Create a legend mapping colors to gear labels
    handles = [plt.Line2D([0], [0], marker='o', color='w',
                          label=str(gear), 
                          markerfacecolor=plt.cm.tab20(idx / max(1, len(unique_gears)-1)), 
                          markersize=8)
               for gear, idx in gear_to_int.items()]
    plt.legend(handles=handles, title='Gear', bbox_to_anchor=(1.05, 1), loc='upper left')
    
    plt.tight_layout()
    plt.show()

def time_passed_during_run(df, time_col='tt_s'):
    """Calculate the total time passed during the run."""
    total_time = df[time_col].max() - df[time_col].min()
    print(f"Total time passed during the run: {total_time}")

def set_start_time_to_zero(df, time_col='tt_s', unix=True):
    """Set the start time to zero."""
    if not unix:
        # convert 2024-06-04T13:36:13.000Z to unix
        df[time_col] = pd.to_datetime(df[time_col]).astype(np.int64) // 10**9
    df[time_col] = df[time_col] - df[time_col].min()
    return df

def get_start_time(df, time_col='tt_s'):
    """Get the start time of the DataFrame."""
    return df[time_col].min()

def time_to_unix(df, time_col='tt_s'):
    """Convert time column to Unix timestamp."""
    df[time_col] = pd.to_datetime(df[time_col]).astype(np.int64) // 10**9
    return df

def unix_to_time(df, time_col='tt_s'):
    """Convert Unix timestamp to datetime."""
    df[time_col] = pd.to_datetime(df[time_col], unit='s')
    return df

def filter_data_by_time(df, time_col='tt_s', start_time=0, end_time=100):
    """Filter the DataFrame to include only data within a specific time range."""
    return df[(df[time_col] >= start_time) & (df[time_col] <= end_time)]

def add_gear_distribution_column(df, gear_distribution_df, time_col='tt_s', lap_col='Lap'):
    """Add a gear distribution column to the DataFrame based on lap and time within lap.
    If gear is None/empty, the previous gear value is carried forward."""
    df = df.copy()
    gear_distribution_df = gear_distribution_df.copy()
    
    # Initialize gear column
    df['Gear'] = None
    
    # Keep track of the last valid gear across all laps
    last_valid_gear = None
    
    # Process each lap in order
    for lap_num in sorted(df[lap_col].unique()):
        if pd.isna(lap_num):
            continue

        # Safely convert lap_num to int if possible
        try:
            lap_num_int = int(lap_num)
        except (ValueError, TypeError):
            continue

        lap_data = df[df[lap_col] == lap_num].copy()
        
        # Get lap start time and convert to relative time within lap
        lap_start_time = lap_data[time_col].min()
        lap_data['lap_relative_time'] = lap_data[time_col] - lap_start_time
        
        # Find corresponding columns in gear distribution for this lap
        time_col_name = f't_lap{lap_num}_nr'
        gear_col_name = f'g_lap{lap_num}_nr'
        
        if time_col_name in gear_distribution_df.columns and gear_col_name in gear_distribution_df.columns:
            # Get gear changes for this lap (remove NaN values)
            gear_times = gear_distribution_df[time_col_name].dropna()
            gear_values = gear_distribution_df[gear_col_name].dropna()
            
            # Convert gear_times to numeric values (seconds)
            gear_times_numeric = []
            for time_val in gear_times:
                if pd.api.types.is_timedelta64_dtype(type(time_val)):
                    gear_times_numeric.append(time_val.total_seconds())
                else:
                    try:
                        gear_times_numeric.append(float(time_val))
                    except (ValueError, TypeError):
                        continue
            
            gear_times_numeric = pd.Series(gear_times_numeric, index=gear_times.index[:len(gear_times_numeric)])
            
            # For each row in this lap, find the appropriate gear
            for idx, row in lap_data.iterrows():
                current_time = row['lap_relative_time']
                
                # Convert current_time to numeric if it's a Timedelta
                if pd.api.types.is_timedelta64_dtype(type(current_time)):
                    current_time = current_time.total_seconds()
                elif hasattr(current_time, 'total_seconds'):
                    current_time = current_time.total_seconds()
                else:
                    current_time = float(current_time)
                
                # Find the most recent gear change before or at current time
                valid_gear_times = gear_times_numeric[gear_times_numeric <= current_time]
                current_gear = None
                
                if not valid_gear_times.empty:
                    # Get index of the most recent gear change
                    gear_idx = valid_gear_times.index[-1]
                    gear_number = gear_values.loc[gear_idx] if gear_idx in gear_values.index else None
                    
                    # Look for H/V indicator in adjacent cells
                    gear_indicator = None
                    try:
                        # Check the cell to the right of the gear number
                        next_col_idx = gear_distribution_df.columns.get_loc(gear_col_name) + 1
                        if next_col_idx < len(gear_distribution_df.columns):
                            next_col = gear_distribution_df.columns[next_col_idx]
                            indicator_value = gear_distribution_df[next_col].loc[gear_idx]
                            if pd.notna(indicator_value) and indicator_value in ['H', 'V']:
                                gear_indicator = indicator_value
                    except (IndexError, KeyError):
                        pass
                    
                    # Combine gear number and indicator
                    if pd.notna(gear_number):
                        if gear_indicator:
                            current_gear = f"{gear_number}{gear_indicator}"
                        else:
                            current_gear = str(gear_number)
                else:
                    # If no gear change found, use first gear value if available
                    if not gear_values.empty:
                        gear_number = gear_values.iloc[0]
                        # Look for H/V indicator for first gear
                        gear_indicator = None
                        try:
                            first_idx = gear_values.index[0]
                            next_col_idx = gear_distribution_df.columns.get_loc(gear_col_name) + 1
                            if next_col_idx < len(gear_distribution_df.columns):
                                next_col = gear_distribution_df.columns[next_col_idx]
                                indicator_value = gear_distribution_df[next_col].loc[first_idx]
                                if pd.notna(indicator_value) and indicator_value in ['H', 'V']:
                                    gear_indicator = indicator_value
                        except (IndexError, KeyError):
                            pass
                        
                        if pd.notna(gear_number):
                            if gear_indicator:
                                current_gear = f"{gear_number}{gear_indicator}"
                            else:
                                current_gear = str(gear_number)
                
                # Set the gear value: use current gear if found, otherwise carry forward last valid gear
                if current_gear is not None:
                    df.at[idx, 'Gear'] = current_gear
                    last_valid_gear = current_gear  # Update last valid gear
                elif last_valid_gear is not None:
                    # Carry forward the previous gear
                    df.at[idx, 'Gear'] = last_valid_gear
        else:
            # If no gear distribution columns exist for this lap, carry forward last gear
            if last_valid_gear is not None:
                for idx in lap_data.index:
                    df.at[idx, 'Gear'] = last_valid_gear
    
    return df

if __name__ == "__main__":
    # Get the directory where this script is located
    script_dir = Path(__file__).parent
    data_dir = script_dir / "Data"
    
    ski_pole_file = "BIA24-3NR_1hz.csv"
    GNSS_file = "BIA24-3_tcx.xlsx"
    
    # Load the data
    data_file = data_dir / "Ski pole data" / ski_pole_file
    df = load_data(str(data_file))

    data_file_gnss = data_dir / "GNSS data" / GNSS_file
    df_gnss = load_xslsx_data_with_row_color_as_column_values(str(data_file_gnss), sheet_name=["NR", "WR"])["NR"]

    print(df_gnss['RowColor'].value_counts())
    gear_distribution_file = data_dir / "Gear distribution 3d.xlsx"
    df_gear_distribution = load_gear_distribution_data(str(gear_distribution_file), sheet_name="BIA24-3")

    # Set start time to zero
    # df = set_start_time_to_zero(df, time_col='tt_s')
    # df_gnss = set_start_time_to_zero(df_gnss, time_col='ns1:Time', unix=False)
    
    time_to_unix(df_gnss, time_col='ns1:Time')
    df_merged = merge_two_dataframes_on_time(df, df_gnss, time_col_1='tt_s', time_col_2='ns1:Time', tolerance=1)
    df_merged = keep_rows_of_colour(df_merged, color_column='RowColor', color_value='FFFFFF')
    df_merged = add_lap_column(df_merged, time_col='tt_s')
    df_merged = add_gear_distribution_column(df_merged, df_gear_distribution, time_col='tt_s', lap_col='Lap')
    #save_df_to_csv(df_merged, '/Users/gustafbjorn/Documents/Chalmers/CAS/TRA300 Digitalization in sports/Project/Project/Playing around with data/merged_data_with_gear.csv')
    print(df_merged['Lap'].value_counts())
    print(df_merged.head())
    print(f"Merged DataFrame shape: {df_merged.shape}")
    plot_lat_lon_map_with_gear_colourmap(df_merged, lat_col='latitude_deg', lon_col='longitude_deg', gear_col='Gear')

    plot_map_from_lat_and_lon_with_lap_colourmap(df_merged, lat_col='latitude_deg', lon_col='longitude_deg')
    plot_lat_lon_difference_map(df_merged, df_merged, lat_col_1='latitude_deg', lon_col_1='longitude_deg', lat_col_2="ns1:LatitudeDegrees", lon_col_2="ns1:LongitudeDegrees")
    # Summarize the data
    # summarize_data(df)
    # summarize_data(df_gnss)

    # Plot histogram for a specific column
    # column_to_plot = 'latitude_deg'  # Update with the column you want to plot
    # plot_histogram(df, column_to_plot)

    # Plot map from latitude and longitude
    # plot_map_from_lat_and_lon_with_time_colourmap(df)
    # plot_map_from_lat_and_lon_with_time_colourmap(df_gnss, lat_col='ns1:LatitudeDegrees', lon_col='ns1:LongitudeDegrees', time_col='ns1:Time')

    df_gnss = keep_rows_of_colour(df_gnss, color_column='RowColor', color_value='FFFFFF')
    plot_and_compare_lat_lon_maps(df_merged, df_merged, lat_col_1='latitude_deg', lon_col_1='longitude_deg', lat_col_2="ns1:LatitudeDegrees", lon_col_2="ns1:LongitudeDegrees", time_col_1='tt_s', time_col_2='ns1:Time')

    # Compare two geographical scatter maps
    start_time = get_start_time(df_gnss, time_col='ns1:Time')
    print(f"GNSS Start time: {datetime.datetime.fromtimestamp(start_time)}")
    start_time_df = get_start_time(df, time_col='tt_s')
    print(f"Ski pole Start time: {datetime.datetime.fromtimestamp(start_time_df)}")
    df = filter_data_by_time(df, time_col='tt_s', start_time=start_time, end_time=start_time + 300)
    df_gnss = filter_data_by_time(df_gnss, time_col='ns1:Time', start_time=start_time, end_time=start_time + 300)
    # plot_and_compare_lat_lon_maps(df, df_gnss, lat_col_1='latitude_deg', lon_col_1='longitude_deg', lat_col_2="ns1:LatitudeDegrees", lon_col_2="ns1:LongitudeDegrees", time_col_1='tt_s', time_col_2='ns1:Time')

    # Calculate time passed during the run
    time_passed_during_run(df, time_col='tt_s')

    # Filter data by time range
    filtered_df = filter_data_by_time(df, time_col='tt_s', start_time=0, end_time=500)
    # plot_map_from_lat_and_lon_with_time_colourmap(filtered_df)

