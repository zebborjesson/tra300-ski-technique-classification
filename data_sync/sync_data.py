import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import re
from pathlib import Path

# ---------- A) Functions for single skier data processing ----------
def load_data(file_path):
    """Load data from a CSV file into a pandas DataFrame."""
    return pd.read_csv(file_path)

def load_xslsx_data(file_path, sheet_name=None):
    """Load data from an Excel file into a pandas DataFrame."""
    return pd.read_excel(file_path, sheet_name)

def load_xslsx_data_with_row_color_as_column_values(file_path, sheet_name=None):
    """Load data from an Excel file and include row colors as column values."""
    
    # Load workbook with openpyxl to get formatting
    wb = load_workbook(file_path)
    df_dict = {}
    
    for sheet in wb.sheetnames:
        if sheet_name and sheet not in (sheet_name if isinstance(sheet_name, list) else [sheet_name]):
            continue
            
        ws = wb[sheet]
        
        # Read data with pandas
        df = pd.read_excel(file_path, sheet_name=sheet)
        
        # Extract row colors
        row_colors = []
        for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (after header)
            cell = ws.cell(row=row_idx, column=1)  # Get first cell of each row
            if cell.fill.start_color.index != '00000000':  # Not default (no fill)
                color = cell.fill.start_color.rgb
                if color:
                    row_colors.append(color)
                else:
                    row_colors.append('FFFFFF')  # Default white
            else:
                row_colors.append('FFFFFF')  # Default white
        
        # Pad row_colors list to match DataFrame length
        while len(row_colors) < len(df):
            row_colors.append('FFFFFF')
        
        # Add row colors as a new column
        df['RowColor'] = row_colors[:len(df)]
        df_dict[sheet] = df
    
    return df_dict

def load_gear_distribution_data(file_path, sheet_name=None):
    """Load gear distribution data from an Excel file into a pandas DataFrame."""
    return pd.read_excel(file_path, sheet_name)

def save_df_to_csv(df, file_path):
    """Save a pandas DataFrame to a CSV file."""
    df.to_csv(file_path, index=False)

def keep_rows_of_colour(df, color_column='RowColor', color_value='FFFF0000'):
    """Keep only the rows of a specific color."""
    return df[df[color_column] == color_value]

def merge_two_dataframes_on_time(df1, df2, time_col_1='tt_s', time_col_2='ns1:Time', tolerance=1):
    """Merge two DataFrames based on time columns with a specified tolerance."""
    df1 = df1.copy()
    df2 = df2.copy()
    df1[time_col_1] = pd.to_datetime(df1[time_col_1], unit='s')
    df2[time_col_2] = pd.to_datetime(df2[time_col_2], unit='s')
    
    merged_df = pd.merge_asof(df1.sort_values(time_col_1), 
                              df2.sort_values(time_col_2), 
                              left_on=time_col_1, 
                              right_on=time_col_2, 
                              tolerance=pd.Timedelta(seconds=tolerance), 
                              direction='nearest')
    return merged_df

def add_lap_column(df, time_col='tt_s', gap_threshold=5):
    """Add a lap column that starts at 1 and increments when there's a gap > threshold."""
    df = df.copy()
    # Calculate time differences between consecutive rows
    time_diff = df[time_col].diff()
    # If time column is datetime, convert gap_threshold to Timedelta
    if np.issubdtype(df[time_col].dtype, np.datetime64):
        gap_mask = time_diff > pd.Timedelta(seconds=gap_threshold)
    else:
        gap_mask = time_diff > gap_threshold
    # Create lap column starting at 1, incrementing at each gap
    df['Lap'] = gap_mask.cumsum() + 1
    return df

def time_to_unix(df, time_col='tt_s'):
    """Convert time column to Unix timestamp."""
    df = df.copy()  # Create explicit copy to avoid SettingWithCopyWarning
    df[time_col] = pd.to_datetime(df[time_col]).astype(np.int64) // 10**9
    return df

def add_gear_distribution_column(df, gear_distribution_df, time_col='tt_s', lap_col='Lap'):
    """Add a gear distribution column to the DataFrame based on lap and time within lap.
    If gear is None/empty, the previous gear value is carried forward."""
    df = df.copy()
    gear_distribution_df = gear_distribution_df.copy()
    
    # Initialize gear column
    df['Gear'] = None
    
    # Keep track of the last valid gear across all laps
    last_valid_gear = None
    
    # Process each lap in order
    for lap_num in sorted(df[lap_col].unique()):
        if pd.isna(lap_num):
            continue

        # Safely convert lap_num to int if possible
        try:
            lap_num_int = int(lap_num)
        except (ValueError, TypeError):
            continue

        lap_data = df[df[lap_col] == lap_num].copy()
        
        # Get lap start time and convert to relative time within lap
        lap_start_time = lap_data[time_col].min()
        lap_data['lap_relative_time'] = lap_data[time_col] - lap_start_time
        
        # Find corresponding columns in gear distribution for this lap
        time_col_name = f't_lap{lap_num}_nr'
        gear_col_name = f'g_lap{lap_num}_nr'
        
        if time_col_name in gear_distribution_df.columns and gear_col_name in gear_distribution_df.columns:
            # Get gear changes for this lap (remove NaN values)
            gear_times = gear_distribution_df[time_col_name].dropna()
            gear_values = gear_distribution_df[gear_col_name].dropna()
            
            # Convert gear_times to numeric values (seconds)
            gear_times_numeric = []
            for time_val in gear_times:
                if pd.api.types.is_timedelta64_dtype(type(time_val)):
                    gear_times_numeric.append(time_val.total_seconds())
                else:
                    try:
                        gear_times_numeric.append(float(time_val))
                    except (ValueError, TypeError):
                        continue
            
            gear_times_numeric = pd.Series(gear_times_numeric, index=gear_times.index[:len(gear_times_numeric)])
            
            # For each row in this lap, find the appropriate gear
            for idx, row in lap_data.iterrows():
                current_time = row['lap_relative_time']
                
                # Convert current_time to numeric if it's a Timedelta
                if pd.api.types.is_timedelta64_dtype(type(current_time)):
                    current_time = current_time.total_seconds()
                elif hasattr(current_time, 'total_seconds'):
                    current_time = current_time.total_seconds()
                else:
                    current_time = float(current_time)
                
                # Find the most recent gear change before or at current time
                valid_gear_times = gear_times_numeric[gear_times_numeric <= current_time]
                current_gear = None
                
                if not valid_gear_times.empty:
                    # Get index of the most recent gear change
                    gear_idx = valid_gear_times.index[-1]
                    gear_number = gear_values.loc[gear_idx] if gear_idx in gear_values.index else None
                    
                    # Look for H/V indicator in adjacent cells
                    gear_indicator = None
                    try:
                        # Check the cell to the right of the gear number
                        next_col_idx = gear_distribution_df.columns.get_loc(gear_col_name) + 1
                        if next_col_idx < len(gear_distribution_df.columns):
                            next_col = gear_distribution_df.columns[next_col_idx]
                            indicator_value = gear_distribution_df[next_col].loc[gear_idx]
                            if pd.notna(indicator_value) and indicator_value in ['H', 'V']:
                                gear_indicator = indicator_value
                    except (IndexError, KeyError):
                        pass
                    
                    # Combine gear number and indicator
                    if pd.notna(gear_number):
                        if gear_indicator:
                            current_gear = f"{gear_number}{gear_indicator}"
                        else:
                            current_gear = str(gear_number)
                else:
                    # If no gear change found, use first gear value if available
                    if not gear_values.empty:
                        gear_number = gear_values.iloc[0]
                        # Look for H/V indicator for first gear
                        gear_indicator = None
                        try:
                            first_idx = gear_values.index[0]
                            next_col_idx = gear_distribution_df.columns.get_loc(gear_col_name) + 1
                            if next_col_idx < len(gear_distribution_df.columns):
                                next_col = gear_distribution_df.columns[next_col_idx]
                                indicator_value = gear_distribution_df[next_col].loc[first_idx]
                                if pd.notna(indicator_value) and indicator_value in ['H', 'V']:
                                    gear_indicator = indicator_value
                        except (IndexError, KeyError):
                            pass
                        
                        if pd.notna(gear_number):
                            if gear_indicator:
                                current_gear = f"{gear_number}{gear_indicator}"
                            else:
                                current_gear = str(gear_number)
                
                # Set the gear value: use current gear if found, otherwise carry forward last valid gear
                if current_gear is not None:
                    df.at[idx, 'Gear'] = current_gear
                    last_valid_gear = current_gear  # Update last valid gear
                elif last_valid_gear is not None:
                    # Carry forward the previous gear
                    df.at[idx, 'Gear'] = last_valid_gear
        else:
            # If no gear distribution columns exist for this lap, carry forward last gear
            if last_valid_gear is not None:
                for idx in lap_data.index:
                    df.at[idx, 'Gear'] = last_valid_gear
    
    return df

# ---------- B) Batch runner ----------

def process_all_skiers_like_single(
    base_dir="Project/Playing around with data/Data",
    gnss_dir="GNSS data",
    pole_dir="Ski pole data",
    gear_xlsx="Gear distribution 3d.xlsx",
    output_dir="outputs",
    color_keep="FFFFFF",
    tolerance_sec=1,
    drop_unmatched=True,
):
    base = Path(base_dir)
    gnss_path = base / gnss_dir
    pole_path = base / pole_dir
    gear_path = base / gear_xlsx
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # 1) load the full gear workbook once (dict of sheets)
    gear_book = load_gear_distribution_data(gear_path, sheet_name=None)  # dict {rider_id: df}
    gear_sheet_names = set(gear_book.keys())

    # 2) find GNSS files for rider IDs (e.g., BIA24-3_tcx.xlsx or "BIA24-14 tcx.xlsx")
    gnss_files_with_underscore = sorted(gnss_path.glob("BIA24-*_tcx.xlsx"))
    gnss_files_with_space = sorted(gnss_path.glob("BIA24-* tcx.xlsx"))
    
    # Combine both patterns, removing duplicates
    all_gnss_files = list(set(gnss_files_with_underscore + gnss_files_with_space))
    
    id_regex = re.compile(r"(BIA24-\d+)")

    riders = {}
    for f in all_gnss_files:
        m = id_regex.search(f.name)
        if m:
            rider_id = m.group(1)
            # Prefer _tcx.xlsx files if both exist
            if rider_id not in riders or "_tcx" in f.name:
                riders[rider_id] = f

    print(f"Found {len(riders)} riders:", sorted(riders.keys()))

    # 3) iterate riders and variants (NR/WR) and run the pipeline
    for rider_id, gnss_file in riders.items():
        if rider_id not in gear_sheet_names:
            print(f"[WARN] No gear sheet for {rider_id}, skipping.")
            continue

        # rider-specific gear sheet
        df_gear_distribution = gear_book[rider_id]

        for variant in ("NR", "WR"):
            # --- Load GNSS with row colors ---
            try:
                gnss_dict = load_xslsx_data_with_row_color_as_column_values(
                    str(gnss_file), sheet_name=[variant]
                )
                df_gnss = gnss_dict[variant]
            except Exception as e:
                print(f"[INFO] {rider_id} {variant}: GNSS sheet missing → skip. ({e})")
                continue

            # --- Keep only rows of the specified color ---
            df_gnss = keep_rows_of_colour(df_gnss, color_column="RowColor", color_value=color_keep)

            # --- Convert GNSS time to unix ---
            if "ns1:Time" not in df_gnss.columns:
                print(f"[WARN] {rider_id} {variant}: 'ns1:Time' not found, skipping.")
                continue
            df_gnss = time_to_unix(df_gnss, time_col="ns1:Time")

            # --- Load ski pole 1 Hz file for this rider+variant ---
            # match common patterns, e.g., BIA24-3NR_1hz.csv / BIA24-3WR_1hz.csv
            pole_candidates = sorted(pole_path.glob(f"{rider_id}{variant}_*1hz*.csv"))
            if not pole_candidates:
                pole_candidates = sorted(pole_path.glob(f"{rider_id}{variant}_*.csv"))
            if not pole_candidates:
                print(f"[INFO] {rider_id} {variant}: no ski pole CSV found, skipping.")
                continue
            pole_file = pole_candidates[0]
            df_pole = load_data(str(pole_file))

            # --- Merge on time ---
            df_merged = merge_two_dataframes_on_time(
                df_pole, df_gnss, time_col_1="tt_s", time_col_2="ns1:Time", tolerance=tolerance_sec
            )

            # keep only rows that matched a GNSS record if wanted
            if drop_unmatched:
                df_merged = df_merged.dropna(subset=["ns1:Time"]).copy()

            # --- Add Lap ---
            df_merged = add_lap_column(df_merged, time_col="tt_s", gap_threshold=5)

            # --- Add gear (with H/V detection) ---
            df_merged = add_gear_distribution_column(
                df_merged, df_gear_distribution, time_col="tt_s", lap_col="Lap"
            )

            # --- Save output per rider+variant ---
            out_file = out_dir / f"{rider_id}_{variant}_merged_with_gear.csv"
            save_df_to_csv(df_merged, str(out_file))
            print(f"[OK] {rider_id} {variant} → {out_file}")


if __name__ == "__main__":
    # Get the directory where this script is located
    script_dir = Path(__file__).parent
    
    # Define paths relative to the script location
    data_dir = script_dir / "Data"
    output_dir = script_dir / "outputs"
    
    process_all_skiers_like_single(
        base_dir=str(data_dir),
        gnss_dir="GNSS data",
        pole_dir="Ski pole data",
        gear_xlsx="Gear distribution 3d_new.xlsx",
        output_dir=str(output_dir),
        color_keep="FFFFFF",
        tolerance_sec=1,
        drop_unmatched=True
    )
